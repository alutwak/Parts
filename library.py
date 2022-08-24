"""Defines the Library class."""
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell

from digikey_api import DigikeyPartError

from part import Part
from parts import Parts
from themes import DEFAULT_THEME
from error import PartError

PART_KEY = "digi_key_part_number"

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium18",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)

MAX_COL_WIDTH = 30


def parse_cell_string(cell):
    """Parse a cell value to the correct type."""
    if not isinstance(cell, str):
        return cell
    if "http" in cell:
        return f'=HYPERLINK("{cell}", "link")'
    try:
        return int(cell)
    except ValueError:
        pass
    try:
        return float(cell)
    except ValueError:
        return cell


class Library:
    """A Parts library (in excel)."""

    def __init__(self, wb_path: str, parts_map_path: str):
        """Initialize stuff."""
        self._wb_path = wb_path
        self._wb = self._get_workbook(wb_path)
        f = open(parts_map_path)
        self._map = yaml.safe_load(f)
        self._parts = Parts()

    def __del__(self):
        """Force the parts to delete immediately."""
        del(self._parts)

    def get_part(self, digikey_part_number: str) -> Part:
        """Get the part from the cache."""
        return self._parts.get_part(digikey_part_number)

    def add_part(self, digikey_part_number: str, stock: int = None):
        """Add the part to the worksheet, and also to the cache, if it's not there."""
        if digikey_part_number not in self._parts:
            self._parts.update_part(digikey_part_number)

        if self.find_part_row(digikey_part_number) is None:
            if stock is None:
                self.add_part_row(digikey_part_number)
            else:
                self.add_part_row(digikey_part_number, stock=stock)

    def add_parts(self, part_numbers: list[str]):
        """Add multiple parts."""
        for part in part_numbers:
            self.add_part(part)

    def add_cached_parts(self):
        """Add all cached parts to the workbook."""
        for part in self._parts:
            if self.find_part_row(part) is None:
                self.add_part_row(part)

    def find_part_row(self, digikey_part_number: str) -> tuple[Cell]:
        """Find the row for the given part."""
        ws = self._get_sheet(self._wb, digikey_part_number)
        if ws is None:
            return None

        table = next(iter(ws.tables.values()))
        key_column_name = self._key_column_name(digikey_part_number)
        col_num = [col.name for col in table.tableColumns].index(key_column_name) + 1
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value == digikey_part_number:
                    return row

    def update_part_row(self, digikey_part_number: str):
        """Update the row for a part."""
        row = self.find_part_row(digikey_part_number)
        if row is None:
            self.add_part_row(digikey_part_number)
        else:
            self._update_part_row(row, digikey_part_number)

    def add_part_row(self, digikey_part_number: str, **kwargs):
        """Add a row for the given part to the correct sheet."""
        ws = self._get_sheet(self._wb, digikey_part_number, create=True)
        row = self._make_part_row(digikey_part_number, **kwargs)
        ws.append(row)

    def resheet_parts(self):
        """Place all of the parts into the correct sheet, according to the map."""
        new_wb = self._create_workbook()
        for ws in self._wb.worksheets:
            col = self._get_part_number_column(ws.title)
            print(f"Part number column: {col}")
            for row in ws.iter_rows(min_row=2, values_only=True):
                part_number = row[col]
                print(f"Part number: {part_number}")
                try:
                    new_ws = self._get_sheet(new_wb, part_number)
                except DigikeyPartError:
                    # If the part number doesn't exist at digikey, fall back to old sheet name
                    new_ws = new_wb.get_sheet_by_name(ws.title)
                new_ws.append(row)
        self._wb = new_wb

    def save(self):
        """Save the worksheet."""
        self._auto_width()
        self._wb.save(self._wb_path)

    def _get_workbook(self, path: str) -> Workbook:
        """Get the workbook."""
        try:
            return load_workbook(path)
        except FileNotFoundError:
            return self._create_workbook()

    def _create_workbook(self) -> Workbook:
        """Create a new workbook with the correct theme and sheets."""
        wb = Workbook()
        del(wb["Sheet"])
        wb.loaded_theme = DEFAULT_THEME
        self._update_sheets(wb)
        return wb

    def _get_part_config(self, digikey_part_number: str) -> (str, dict):
        """Get the sheet configuration on which the part belongs."""
        part = self.get_part(digikey_part_number)
        for name, conf in self._map.items():
            if "Taxonomies" in conf and part.taxonomy in conf['Taxonomies']:
                return (name, conf)
        return ("Unsorted", self._map["Unsorted"])

    def _get_sheet(self, wb: Workbook, digikey_part_number: str, create: bool = False) -> Worksheet:
        """Get the sheet on which the part belongs."""
        name, config = self._get_part_config(digikey_part_number)
        try:
            return wb.get_sheet_by_name(name)
        except KeyError:
            if create:
                # Create a new sheet and the table that goes on it
                return self._make_sheet(wb, name, config)

    def _get_sheet_config(self, sheet_name: str) -> dict:
        """Return the configuration for the given sheet name."""
        return self._map[sheet_name]

    def _get_part_number_column(self, sheet_name: str) -> int:
        """Return the column of the digikey part number for the given sheet."""
        col = 0
        for section, config in self._get_sheet_config(sheet_name).items():
            if section == "Taxonomies":
                continue
            for key in config.keys():
                if key == "digi_key_part_number":
                    return col
                col += 1
        raise PartError(f"Unable to get part number column for {sheet_name} sheet")

    def _update_sheets(self, wb: Workbook):
        """Ensure that all sheets defined in the part map exist in the given workbook."""
        for name, config in self._map.items():
            if "Taxonomies" in config and name not in wb.sheetnames:
                self._make_sheet(wb, name, config)

    def _sheet_header(self, config) -> list[str]:
        """Define the sheet header for the given config."""
        sections = [x for x in config.keys() if x != "Taxonomies"]
        header = []
        for section in sections:
            try:
                for col in config[section].values():
                    header.append(col)
            except AttributeError:
                raise PartError(f"{section} section is ill-formatted")
        return header

    def _make_sheet(self, wb: Workbook, name: str, config: dict) -> Worksheet:
        """Make the sheet and table defined by name and config."""
        ntables = len([sheet for sheet in wb.sheetnames if sheet != "Sheet"]) + 1
        ws = wb.create_sheet(name)
        try:
            header = self._sheet_header(config)
        except PartError as e:
            raise PartError(f"Error configuring {name} sheet: {e}")
        ws.append(header)
        end_row = get_column_letter(len(header))
        ref = f"A1:{end_row}1048576"
        table = Table(displayName=f"Table{ntables}", ref=ref, tableStyleInfo=TABLE_STYLE)
        table._initialise_columns()
        for col, head in zip(table.tableColumns, header):
            col.name = head
        ws.add_table(table)
        ws.freeze_panes = "B2"
        return ws

    def _key_column_name(self, digikey_part_number: str) -> str:
        """Get the key column name for the given config."""
        name, config = self._get_part_config(digikey_part_number)
        sections = [x for x in config.keys() if x != "Taxonomies"]
        for section in sections:
            for key, col_name in config[section].items():
                if key == PART_KEY:
                    return col_name
        raise KeyError(f"Failed to find {PART_KEY} parameter in the {name} section")

    def _make_part_row(self, digikey_part_number: str, **kwargs) -> list[str]:
        """Make a row from the given part number.

        You can use kwargs to define starting values for the '$' parameters
        """
        part = self._parts.get_part(digikey_part_number)
        name, config = self._get_part_config(digikey_part_number)
        row = []
        for section, params in config.items():
            if section == "Taxonomies":
                continue
            for key in params.keys():
                try:
                    if key[0] == "$":
                        if key[1:] in kwargs:
                            row.append(kwargs[key[1:]])
                        else:
                            row.append("")
                    else:
                        row.append(parse_cell_string(part[key]))
                except PartError:
                    row.append("")
        return row

    def _update_part_row(self, row: tuple[Cell], digikey_part_number: str):
        """Update the cells in the row."""
        part = self._parts.update_part(digikey_part_number)
        name, config = self._get_part_config(digikey_part_number)

        params = []
        for section, params in config.items():
            if section == "Taxonomies":
                continue
            params.extend(params.keys())

        for cell, param in zip(row, params):
            if param[0] == "$" or (isinstance(cell.value, str) and cell.value[0] == "'"):
                continue
            else:
                try:
                    cell.value = parse_cell_string(part[param])
                except PartError:
                    pass

    def _auto_width(self):
        """Set column widths automatically."""
        for ws in self._wb.worksheets:
            for c in range(ws.max_column):
                rows = list(ws.rows)
                header_width = len(rows[0][c].value) + 3
                width = header_width
                for row in rows[1:]:
                    w = len(str(row[c].value))
                    width = max(width, w)
                if width > MAX_COL_WIDTH:
                    width = header_width

                letter = get_column_letter(c + 1)
                ws.column_dimensions[letter].width = width
